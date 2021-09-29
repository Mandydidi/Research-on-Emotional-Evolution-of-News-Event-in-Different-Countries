#coding:utf-8
from pythainlp import word_tokenize
from tqdm import tqdm_notebook
from pythainlp.ulmfit import process_thai
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
import time

'''
#训练文本和标签放入all_df中
with open("D://cs//python//project//train.txt", 'r', encoding='utf-8') as f:
    texts = [line.strip() for line in f.readlines()]
with open("D://cs//python//project//train_label.txt", 'r', encoding='utf-8') as f:
    categories = [line.strip() for line in f.readlines()]
all_df = pd.DataFrame({"category":categories, "texts":texts})
all_df.to_excel('D://cs//python//project//all_df.xlsx', index=False)
print('训练文本和标签存入excel文件成功')
with open("D://cs//python//project//valid.txt", 'r', encoding='utf-8') as f:
    texts = [line.strip() for line in f.readlines()]
with open("D://cs//python//project//valid_label.txt", 'r', encoding='utf-8') as f:
    categories = [line.strip() for line in f.readlines()]
valid_df = pd.DataFrame({"category":categories, "texts":texts})
valid_df.to_excel('D://cs//python//project//valid_df.xlsx', index=False)
print('确定集文本和标签存入excel成功')

#获取测试集,后续修改路径
with open("D://cs//python//project//test.txt", 'r', encoding='utf-8') as f:
    texts = [line.strip() for line in f.readlines()]
test_df = pd.DataFrame({"category":'test', "texts":texts})

#加载数据
all_df["processed"] = all_df.texts.map(lambda x: "|".join(process_thai(x)))  #分词后用竖线连接
all_df["wc"] = all_df.processed.map(lambda x: len(x.split("|")))   #分词后词数
all_df["uwc"] = all_df.processed.map(lambda x: len(set(x.split("|"))))  #分词后不重复词数
all_df.to_excel('D://cs//python//project//all_df.xlsx', index=False, encoding='utf-8')
print('更新后训练集存入excel成功')
valid_df["processed"] = valid_df.texts.map(lambda x: "|".join(process_thai(x)))
valid_df["wc"] = valid_df.processed.map(lambda x: len(x.split("|")))
valid_df["uwc"] = valid_df.processed.map(lambda x: len(set(x.split("|"))))
valid_df.to_excel('D://cs//python//project//valid_df.xlsx', index=False, encoding='utf-8')
print('更新后确定集存入excel成功')
test_df["processed"] = test_df.texts.map(lambda x: "|".join(process_thai(x)))
test_df["wc"] = test_df.processed.map(lambda x: len(x.split("|")))
test_df["uwc"] = test_df.processed.map(lambda x: len(set(x.split("|"))))

#数据预准备
train_df = all_df.copy()  #训练集
y_train = all_df["category"]
y_valid = valid_df["category"]
tfidf = TfidfVectorizer(tokenizer=process_thai, ngram_range=(1,2), min_df=20, sublinear_tf=True)
tfidf_fit = tfidf.fit(all_df["texts"])
text_train = tfidf_fit.transform(train_df["texts"])
text_valid = tfidf_fit.transform(valid_df["texts"])
text_test = tfidf_fit.transform(test_df["texts"])

#计数
scaler = StandardScaler()
scaler_fit = scaler.fit(all_df[["wc","uwc"]].astype(float))
num_train = scaler_fit.transform(train_df[["wc","uwc"]].astype(float))
num_valid = scaler_fit.transform(valid_df[["wc","uwc"]].astype(float))
num_test = scaler_fit.transform(test_df[["wc","uwc"]].astype(float))

#合并文字和字数统计功能
X_train = np.concatenate([num_train,text_train.toarray()], axis=1)
X_valid = np.concatenate([num_valid,text_valid.toarray()],axis=1)
X_test = np.concatenate([num_test,text_test.toarray()], axis=1)

print('数据准备完成')

#匹配模型
model = LogisticRegression(C=2., penalty="l2", solver="liblinear", dual=False, multi_class="ovr")
model.fit(X_train, y_train)
print('模型得分为：'+str(float(model.score(X_valid,y_valid))))

#预测测试集
probs = model.predict_proba(X_test)
probs_df = pd.DataFrame(probs)
probs_df.columns = model.classes_
probs_df["preds"] = model.predict(X_test)
probs_df["texts"] = test_df.texts
probs_df["processed"] = test_df.processed
probs_df["wc"] = test_df.wc
probs_df["uwc"] = test_df.uwc
probs_df.to_excel('D://cs//python//project//test_pred.xlsx', index=False, encoding='utf-8')
print('预测测试集并存入excel成功')
'''

#从准备好的训练集和验证集读出数据
all_df = pd.read_excel('D://cs//python//project//all_df.xlsx', encoding='utf-8')
valid_df = pd.read_excel('D://cs//python//project//valid_df.xlsx', encoding='utf-8')
print('训练集和验证集获取成功')
#数据准备
train_df = all_df.copy()  #训练集
y_train = all_df["category"]
y_valid = valid_df["category"]
tfidf = TfidfVectorizer(tokenizer=process_thai, ngram_range=(1,2), min_df=20, sublinear_tf=True)
tfidf_fit = tfidf.fit(all_df["texts"])
text_train = tfidf_fit.transform(train_df["texts"])
text_valid = tfidf_fit.transform(valid_df["texts"])
#计数
scaler = StandardScaler()
scaler_fit = scaler.fit(all_df[["wc","uwc"]].astype(float))
num_train = scaler_fit.transform(train_df[["wc","uwc"]].astype(float))
num_valid = scaler_fit.transform(valid_df[["wc","uwc"]].astype(float))
#合并文字和字数统计功能
X_train = np.concatenate([num_train,text_train.toarray()], axis=1)
X_valid = np.concatenate([num_valid,text_valid.toarray()],axis=1)
print('数据准备完成')
#匹配模型
model = LogisticRegression(C=2., penalty="l2", solver="liblinear", dual=False, multi_class="ovr")
model.fit(X_train, y_train)
print('模型得分为：'+str(float(model.score(X_valid, y_valid))))
#预测数据
na_list = ['news3_03-09.xlsx', 'news3_03-10.xlsx', 'news3_03-11.xlsx', 'news3_03-12.xlsx', 'news3_03-13.xlsx',
            'news4_03-02.xlsx', 'news4_03-03.xlsx', 'news4_03-04.xlsx', 'news4_03-05.xlsx', 'news4_03-06.xlsx', 'news4_03-07.xlsx', 'news4_03-09.xlsx', 'news4_03-10.xlsx', 'news4_03-11.xlsx', 'news4_03-12.xlsx', 'news4_03-13.xlsx']
#一个文件中一列全部预测
for name in na_list:
    path = 'D://cs//python//project//'+str(name)
    wb = load_workbook(path)
    ws = wb['Sheet']
    row = ws.max_row
    col = ws.max_column
    texts = []
    for i in range(1, row+1):
        texts.append(str(ws.cell(row=i, column=col).value))
    test_df = pd.DataFrame({"category": 'test', "texts": texts})
    test_df["processed"] = test_df.texts.map(lambda x: "|".join(process_thai(x)))
    test_df["wc"] = test_df.processed.map(lambda x: len(x.split("|")))
    test_df["uwc"] = test_df.processed.map(lambda x: len(set(x.split("|"))))
    text_test = tfidf_fit.transform(test_df["texts"])
    num_test = scaler_fit.transform(test_df[["wc", "uwc"]].astype(float))
    X_test = np.concatenate([num_test, text_test.toarray()], axis=1)
    print(str(name)+'测试集准备成功')
    # 预测测试集
    probs = model.predict_proba(X_test)
    probs_df = pd.DataFrame(probs)
    probs_df.columns = model.classes_
    probs_df["preds"] = model.predict(X_test)
    print('预测完成')
    for i in range(1, row+1):
        ws.cell(row=i, column=col+1).value = str(probs_df["preds"][i-1])
    wb.save(path)
    time.sleep(1)
    print('写入成功')
    print('------------------')
