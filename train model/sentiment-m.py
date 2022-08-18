import fastText.FastText as fasttext
from openpyxl import load_workbook
import time

#训练模型
'''
model = fasttext.train_supervised('D://cs//python//project//label-malaya.txt')  #有监督训练，给定文本每一行以__label__标签名开始 后面为该类别下词语
model.save_model('D://cs//python//project//model-senti-m.bin')  #保存模型
'''
#引用模型预测情感
def senti_pre(path):
    wb = load_workbook(path)
    ws = wb['Sheet']
    print('加载excel文件成功'+str(path.split('//')[-1]))
    row = ws.max_row
    col = ws.max_column
    model = fasttext.load_model('D://cs//python//project//model-senti-m.bin')  #加载模型
    print('加载模型成功')
    for i in range(1, row+1):
        print('第'+str(i)+'行：')
        texts = str(ws.cell(row=i, column=col).value)
        res = str(model.predict(texts)[0][0]).replace('__label__', '')
        print('情感倾向为：'+str(res))
        ws.cell(row=i, column=col+1).value = str(res)
        time.sleep(1)
        print('写入成功')
        print('----------------')
    wb.save(path)
    print('保存成功')

name_list = ['news1_03-01.xlsx', 'news1_03-02.xlsx', 'news1_03-03.xlsx', 'news1_03-04.xlsx', 'news1_03-05.xlsx', 'news1_03-06.xlsx', 'news1_03-07.xlsx', 'news1_03-09.xlsx', 'news1_03-10.xlsx', 'news1_03-11.xlsx', 'news1_03-12.xlsx', 'news1_03-13.xlsx',
             'news2_03-01.xlsx', 'news2_03-02.xlsx', 'news2_03-03.xlsx', 'news2_03-04.xlsx', 'news2_03-05.xlsx', 'news2_03-06.xlsx', 'news2_03-07.xlsx', 'news2_03-09.xlsx', 'news2_03-10.xlsx', 'news2_03-11.xlsx', 'news2_03-12.xlsx', 'news2_03-13.xlsx', ]
for name in name_list:
    path = 'D://cs//python//project//'+name
    senti_pre(path)
    time.sleep(1)
    print('***************')
