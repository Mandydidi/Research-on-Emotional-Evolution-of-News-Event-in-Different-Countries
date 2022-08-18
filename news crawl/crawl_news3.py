from fake_useragent import UserAgent
import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import datetime


ua = UserAgent()
def visit_mainpage():
    url = 'https://www.dailynews.co.th/'
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Host': 'www.dailynews.co.th',
        'User-Agent': ua.random,
    }
    response = requests.get(url=url, headers=headers, timeout=500)
    if response.status_code !=200:
        print('访问主页失败，错误码：'+str(response.status_code))
    else:
        print('访问主页成功')
        cookie_dict = {}
        for key, value in response.cookies.items():
            cookie_dict[key] = value
        print('访问主页获取cookie成功')
        f = open('D://cs//python//project//crawl3_cookie.txt', 'w+')
        f.write(str(cookie_dict))
        f.close()
        print('文件写入cookie成功')
        response = response.text
        html = etree.HTML(response)
        link_list = html.xpath('//a//@href')
        print('获取主页面所有链接成功')
        kind_list = ['article', 'economic', 'education', 'entertainment', 'foreign', 'it', 'politics', 'sports', 'women', 'crime']
        link_lists = [[], [], [], [], [], [], [], [], [], []]
        for link in link_list:
            #去除无关链接
            if ('http' in link) or ('video' in link):
                continue
            else:
                classify_link(link, kind_list, link_lists)
        wb = load_workbook('D://cs//python//project//crawl3_link.xlsx')
        ws = wb['Sheet1']
        for i in range(0, len(link_lists)):
            print('列号为:'+str(i+1))
            links = list(set(link_lists[i]))#去重
            for j in range(0, len(links)):
                print('第'+str(j+2)+'行为:'+str(links[j]))
                ws.cell(2+j, i+1, links[j])
        wb.save('D://cs//python//project//crawl3_link.xlsx')
        wb.close()
    print('写入链接成功')


#链接分类
def classify_link(link, kind_list, link_lists):
    for i in range(0, len(kind_list)):
        kind = kind_list[i]
        kind_s = '/'+str(kind)+'/'
        if kind_s in link:#属于某一类
            link = link.replace(kind_s, '')
            if len(link) == 6:#为六位数字串
                link_lists[i].append(link)
            break#不为六位数字串或者已经添加完成，直接退出




#爬取单篇新闻
def crawl_text():
    kind_list = ['article', 'economic', 'education', 'entertainment', 'foreign', 'it', 'politics', 'sports', 'women', 'crime']
    wb1 = load_workbook('D://cs//python//project//crawl3_link.xlsx')
    ws1 = wb1['Sheet1']
    # 创建一个新的excel
    wb2 = Workbook()
    # 指定wb2为当前操作表
    ws2 = wb2.active
    i = 0#初始化属性值
    for col in ws1.columns:#迭代器获取列，col为元组
        for j in range(1, len(col)):
            num = str(col[j].value)#获得单元格值即六位数字串
            if len(num) > 6:#不为六位数字串，读取下一行
                continue
            if (num == '') or (len(num) == 0):#一列已经读完，换下一列
                break
            if len(num) == 6:#有效的六位数字码
                link = 'https://www.dailynews.co.th/'+str(kind_list[i])+'/'+str(num)
                #获得内容列表[link,title,desc,time,text]
                content = visit_page(link)
                if len(content) == 0:#整页内容为空
                    continue
                else:
                    ws2.append(content)
                    wb2.save('D://cs//python//project//news3_{day}.xlsx'.format(day=datetime.datetime.now().strftime("%m-%d")))
                    time.sleep(1.5)
                    print('写入该行内容成功')
                    print('-----------------')
                    time.sleep(2)
        #属性值加一，获取下一列
        i = i+1
    wb1.close()
    wb2.close()
    print(datetime.datetime.now().strftime("%m-%d %H:%M:%S")+'目前文件中单篇新闻写入成功')


#获取单页链接的内容列表
def visit_page(link):
    url = link
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Host': 'www.dailynews.co.th',
        'User-Agent': ua.random,
    }
    response = requests.get(url=url, headers=headers, timeout=500)
    if response.status_code != 200:
        print('访问单篇新闻页面失败，链接为： '+url)
        return []
    else:
        print('访问单篇新闻页面成功')
        html = etree.HTML(response.text)
        title = html.xpath('//h1[@class="title"]//text()')
        date = html.xpath('//p[@class="desc"]//span//text()')
        text = html.xpath('//section[@class="article-detail"]//div//text()')
        if (len(title) == 0) or (title == []) or (title == [None]):#空页面
            print('无标题空页面')
            return []
        content = [url, clean1(title), clean1(date), clean1(text)]
        print('获取该页面内容为:'+str(content))
        return content


#清除/r/n空格,返回字符串
def clean1(v):
    for i in range(0, len(v)):
        v[i] = v[i].replace('/r', '')
        v[i] = v[i].replace('/n', '')
        v[i] = v[i].strip()
    v = list(filter(None, v))#删去列表中空字符串
    return ''.join(v)


visit_mainpage()
crawl_text()






