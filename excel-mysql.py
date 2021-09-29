# -*- coding:utf-8 -*-
import pymysql
import time
from openpyxl import load_workbook

conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='news', charset='utf8')  #连接本地数据库
cursor = conn.cursor()  #获得游标
#'news1_02-29', 'news1_03-01', 'news1_03-02','news1_03-03', 'news1_03-04', 'news1_03-05', 'news1_03-06', 'news1_03-07', 'news1_03-09', 'news1_03-10', 'news1_03-11', 'news1_03-12', 'news1_03-13',
# 'news2_03-01', 'news2_03-02', 'news2_03-03', 'news2_03-04', 'news2_03-05', 'news2_03-06', 'news2_03-07', 'news2_03-09', 'news2_03-10', 'news2_03-11', 'news2_03-12', 'news2_03-13',
#             'news3_02-24', 'news3_02-25', 'news3_02-26', 'news3_02-27', 'news3_02-28', 'news3_02-29', 'news3_03-01', 'news3_03-02', 'news3_03-03', 'news3_03-04', 'news3_03-05', 'news3_03-06',
#              'news3_03-07',
name_list = [
             'news3_03-09', 'news3_03-10', 'news3_03-11', 'news3_03-12', 'news3_03-13',
             'news4_02-28', 'news4_02-29', 'news4_03-01', 'news4_03-02', 'news4_03-03', 'news4_03-04', 'news4_03-05', 'news4_03-06', 'news4_03-07', 'news4_03-09', 'news4_03-10', 'news4_03-11', 'news4_03-12', 'news4_03-13']
#遍历每一个文件名
for name in name_list:
    #创建表
    name_table = name.replace('-', '_')
    create_table = "CREATE TABLE "+str(name_table)+" (" \
                                             "Url varchar(256), " \
                                             "Title varchar(256), " \
                                             "Date varchar(256), " \
                                             "Text text(65535), " \
                                             "Sentiment varchar(256)" \
                                             ")ENGINE=innodb DEFAULT CHARSET=utf8;"
    # 执行SQL语句
    cursor.execute(create_table)
    print('创建名为'+str(name_table)+'表成功')
    #注入数据
    path = 'D://cs//python//project//'+str(name)+'.xlsx'
    wb = load_workbook(path, read_only=True)
    ws = wb['Sheet']
    row = ws.max_row
    col = ws.max_column
    data_list = []
    for i in range(1, row+1):
        line_list = []  #一行以列表形式存储
        for j in range(1, col+1):
            item = str(ws.cell(row=i, column=j).value)  #数据预处理
            item = item.replace('\n', '')
            item = item.replace('\r', '')
            item = item.strip()
            line_list.append(item)
        data_list.append(tuple(line_list))  #一行数据以元组方式添加进列表
        print(str(i)+'行数据加入列表成功')
    print(str(name)+'文件读取成功')
    wb.close()
    insert_lines = "insert into "+str(name_table)+"(Url,Title,Date,Text,Sentiment) values(%s,%s,%s,%s,%s);"
    # 拼接并执行sql语句
    cursor.executemany(insert_lines, data_list)
    # 涉及写操作要注意提交
    conn.commit()
    time.sleep(2)   #写入等待
    print(str(name_table)+'表写入成功')
    print('------------------------')

#关闭连接
cursor.close()
conn.close()
